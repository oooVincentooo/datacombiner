# -*- coding: utf-8 -*-
"""
Created on Mon Jan 10 11:22:33 2022

@author: vpreemen
"""


from openpyxl import load_workbook


class data_xls:

    def __init__(self, recipe):
  
        self.recipe=recipe
        self.datasize=0
        
        self.sheet=self.recipe['Import (excel)']['general']['sheet']['value']
        self.column=self.recipe['Import (excel)']['general']['column']['value']        
        
       
    def data_format(self,section):
        
        self.active=self.recipe['Import (excel)'][section]['active']['value']
        self.search=self.recipe['Import (excel)'][section]['search']['value']
        self.offset=self.recipe['Import (excel)'][section]['offset']['value']
        self.location=self.recipe['Import (excel)'][section]['location']['value']
        self.rows=self.recipe['Import (excel)'][section]['rows']['value']
        

    def data_open(self, path):
        #https://www.geeksforgeeks.org/how-to-iterate-through-excel-rows-in-python/
        #https://openpyxl.readthedocs.io/en/stable/usage.html
        
        self.wb = load_workbook(path, data_only=True, read_only=True)

    def data_sheet(self,sheet):
        self.worksheet= self.wb[sheet]
        
    def data_sheets(self):
        sheets=[]

        if self.sheet=='1':
            g_sheet=self.wb.sheetnames
            sheets=[g_sheet[0]]
        
        elif self.sheet=='all':
            
            
            for sheet in self.wb:
                
                column=self.wb[sheet.title].max_column
                row=self.wb[sheet.title].max_row
                
                if (column+row)>2:
                    sheets.append(sheet.title)
        
        else:
            sheets=[self.sheet]
        
        return sheets           
        

    def data_close(self):
        
        self.wb.close()


    def data_search(self):
        
        self.datasize=self.offset + self.rows
        
        if self.datasize>=self.worksheet.max_row:
           self.datasize=self.worksheet.max_row

 
        if self.location=="top":

           self.start=0       
           result=self.data_find()

        else:
            self.start=self.worksheet.max_row-self.datasize      
            result=self.data_find()
             
        result=[result[0], self.data_array(result[0]), self.search,  result[1]]  
       
        return result        

         
    def data_find(self):
        result=[1, False] 
        for n in range(self.datasize):
            
            #self.start=self.worksheet.max_row-self.datasize

            for m, cell in enumerate(self.worksheet[self.start+n+1]):
            
                if m==self.column-1:


                    if self.search==str(cell.value): 
                        
                        result=[1+self.start+n+self.offset, True] # The string is found
                        break
 
           
                elif m>self.column-1:
                    break
                
            if result[1]:
                break

        return result

    
    def data_header(self):
     
        self.data_format("header") 
 
        if self.active==1:
            if self.search.lower()=='none' or len(self.search)==0:
                
                n=self.offset
                result=[n,self.data_array(n),'',True]
                                
            else:    
                result=self.data_search()
 
        else:
                result=[1,['Column'],'',False]        
             
        return result             


    def data_start(self):
     
        self.data_format("start data") 

 
        if self.active==1:
            if self.search.lower()=='none' or len(self.search)==0:
                
                n=self.offset               
                result=[n,self.data_array(n),'',True]
                           
            else:    
                result=self.data_search()
 
        else:
                result=[1,[''],'',False]        
             
        return result  


    def data_end(self):
     
        self.data_format("end data") 
 
        if self.active==1:
            if self.search.lower()=='none' or len(self.search)==0:
                
                n=self.offset
                result=[n,self.data_array(n),'',True]
                                
            else:    
                result=self.data_search()
 
        else:
                result=[self.worksheet.max_row,[''],'',False]        
             
        return result  

 
    def data_array(self,n):
        data=[]
        for cell in self.worksheet[n]:
            data.append(str(cell.value))
        
        return data

    
